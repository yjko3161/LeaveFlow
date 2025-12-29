import openpyxl
from openpyxl.drawing.image import Image as XLImage
from app.models.approval import ApprovalRequest
from app import db
import os
from flask import current_app

class ExportService:
    def generate_document(self, request_id):
        req = ApprovalRequest.query.get(request_id)
        if not req or req.status != "APPROVED":
            return None, "Not approved"
            
        # Select Template based on doc_type
        # For MVP, we assume a single template or hardcoded generation
        # In a real app, load from export_templates table.
        
        # Let's create a basic workbook dynamically for MVP proof
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Document"
        
        # Headers
        ws['A1'] = "FlexLite ERP Document"
        ws['A2'] = f"Type: {req.doc_type}"
        ws['A3'] = f"ID: {req.doc_id}"
        ws['A4'] = f"Requester: {req.requester.name}"
        ws['A5'] = f"Date: {req.created_at.strftime('%Y-%m-%d')}"
        
        # Specific Content
        row = 7
        if req.doc_type == "LEAVE":
            from app.models.leave import LeaveRequest
            doc = LeaveRequest.query.get(req.doc_id)
            ws[f'A{row}'] = "Leave Type"; ws[f'B{row}'] = doc.leave_type.value
            row+=1
            ws[f'A{row}'] = "Period"; ws[f'B{row}'] = f"{doc.start_date} ~ {doc.end_date} ({doc.days} days)"
            row+=1
            ws[f'A{row}'] = "Reason"; ws[f'B{row}'] = doc.reason
            
        elif req.doc_type == "EXPENSE":
            from app.models.expense import ExpenseReport
            doc = ExpenseReport.query.get(req.doc_id)
            ws[f'A{row}'] = "Title"; ws[f'B{row}'] = doc.title
            row+=1
            ws[f'A{row}'] = "Total Amount"; ws[f'B{row}'] = doc.total_amount
            row+=2
            
            # Table Header
            ws[f'A{row}'] = "Date"
            ws[f'B{row}'] = "Merchant"
            ws[f'C{row}'] = "Amount"
            ws[f'D{row}'] = "Category"
            row+=1
            
            for item in doc.items:
                ws[f'A{row}'] = str(item.usage_date)
                ws[f'B{row}'] = item.merchant
                ws[f'C{row}'] = item.amount
                ws[f'D{row}'] = item.category.name
                row+=1

        # Signatures (Approvers)
        # In real template, we would find the cell named "SIGN_APPROVER_1" etc.
        # Here we just append at bottom
        row += 3
        ws[f'A{row}'] = "Approvals"
        row += 1
        
        for action in req.actions:
            if action.action == "APPROVE":
                ws[f'A{row}'] = f"{action.actor.name} (Approved at {action.created_at.strftime('%Y-%m-%d')})"
                # If we had an image signature:
                # img = XLImage(action.actor.signature.image_path)
                # ws.add_image(img, f'B{row}')
                row += 1
                
        # Save
        filename = f"{req.doc_type}_{req.doc_id}_export.xlsx"
        path = os.path.join(current_app.config['UPLOAD_DIR'], filename)
        wb.save(path)
        
        return path, filename
